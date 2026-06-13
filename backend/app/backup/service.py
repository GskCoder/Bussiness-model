"""Backup service — export and import all data."""

import io
import json
from datetime import date, datetime
from decimal import Decimal
from sqlalchemy.orm import Session

from app.auth.models import User
from app.products.models import Product, Category
from app.customers.models import Customer
from app.sales.models import Sale, SaleItem
from app.inventory.models import InventoryTransaction
from app.expenses.models import Expense, ExpenseCategory
from app.suppliers.models import Supplier
from app.purchases.models import Purchase, PurchaseItem
from app.settings.models import ShopSettings


class CustomEncoder(json.JSONEncoder):
    """JSON encoder that handles dates, Decimals, and enums."""
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if isinstance(obj, Decimal):
            return float(obj)
        if hasattr(obj, 'value'):
            return obj.value
        return super().default(obj)


def _serialize_table(db: Session, model, exclude_cols=None) -> list[dict]:
    """Serialize all rows from a table to list of dicts."""
    exclude = set(exclude_cols or [])
    rows = db.query(model).all()
    result = []
    for row in rows:
        d = {}
        for col in row.__table__.columns:
            if col.name in exclude:
                continue
            val = getattr(row, col.name)
            d[col.name] = val
        result.append(d)
    return result


def export_json(db: Session) -> io.BytesIO:
    """Export all data as a single JSON file."""
    data = {
        "export_date": datetime.now().isoformat(),
        "version": "1.0",
        "categories": _serialize_table(db, Category),
        "products": _serialize_table(db, Product),
        "customers": _serialize_table(db, Customer),
        "suppliers": _serialize_table(db, Supplier),
        "sales": _serialize_table(db, Sale),
        "sale_items": _serialize_table(db, SaleItem),
        "purchases": _serialize_table(db, Purchase),
        "purchase_items": _serialize_table(db, PurchaseItem),
        "inventory_transactions": _serialize_table(db, InventoryTransaction),
        "expense_categories": _serialize_table(db, ExpenseCategory),
        "expenses": _serialize_table(db, Expense),
        "settings": _serialize_table(db, ShopSettings),
        "users": _serialize_table(db, User, exclude_cols=["hashed_password"]),
    }

    buffer = io.BytesIO()
    buffer.write(json.dumps(data, cls=CustomEncoder, indent=2).encode("utf-8"))
    buffer.seek(0)
    return buffer


def export_excel(db: Session) -> io.BytesIO:
    """Export all data as an Excel workbook with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    tables = [
        ("Categories", Category),
        ("Products", Product),
        ("Customers", Customer),
        ("Suppliers", Supplier),
        ("Sales", Sale),
        ("Sale Items", SaleItem),
        ("Purchases", Purchase),
        ("Purchase Items", PurchaseItem),
        ("Inventory Ledger", InventoryTransaction),
        ("Expense Categories", ExpenseCategory),
        ("Expenses", Expense),
    ]

    for sheet_name, model in tables:
        ws = wb.create_sheet(title=sheet_name)
        rows = _serialize_table(db, model)

        if not rows:
            ws.cell(row=1, column=1, value="No data")
            continue

        # Headers
        headers = list(rows[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, key in enumerate(headers, 1):
                val = row_data.get(key)
                if isinstance(val, (datetime, date)):
                    val = val.isoformat()
                elif isinstance(val, Decimal):
                    val = float(val)
                elif hasattr(val, 'value'):
                    val = val.value
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-width
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header) + 2
            for row_data in rows[:50]:
                val = str(row_data.get(header, ""))
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 35)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_csv(db: Session, table_name: str) -> io.BytesIO:
    """Export a single table as CSV."""
    import csv

    model_map = {
        "products": Product,
        "categories": Category,
        "customers": Customer,
        "suppliers": Supplier,
        "sales": Sale,
        "sale_items": SaleItem,
        "purchases": Purchase,
        "inventory": InventoryTransaction,
        "expense_categories": ExpenseCategory,
        "expenses": Expense,
    }

    model = model_map.get(table_name)
    if not model:
        raise ValueError(f"Unknown table: {table_name}")

    rows = _serialize_table(db, model)

    buffer = io.StringIO()
    if rows:
        writer = csv.DictWriter(buffer, fieldnames=rows[0].keys())
        writer.writeheader()
        for row in rows:
            # Convert non-serializable types
            clean = {}
            for k, v in row.items():
                if isinstance(v, (datetime, date)):
                    clean[k] = v.isoformat()
                elif isinstance(v, Decimal):
                    clean[k] = float(v)
                elif hasattr(v, 'value'):
                    clean[k] = v.value
                else:
                    clean[k] = v
            writer.writerow(clean)

    result = io.BytesIO(buffer.getvalue().encode("utf-8"))
    result.seek(0)
    return result


def import_json(db: Session, data: dict) -> dict:
    """Import data from a JSON backup — restores categories, products, customers, etc.
    
    This is an additive import — does not delete existing data.
    Skips records that would cause unique constraint violations.
    """
    stats = {}

    # Import categories
    imported = 0
    for cat_data in data.get("categories", []):
        existing = db.query(Category).filter(Category.name == cat_data.get("name")).first()
        if not existing:
            db.add(Category(name=cat_data["name"], description=cat_data.get("description", "")))
            imported += 1
    stats["categories"] = imported

    # Import expense categories
    imported = 0
    for cat_data in data.get("expense_categories", []):
        existing = db.query(ExpenseCategory).filter(ExpenseCategory.name == cat_data.get("name")).first()
        if not existing:
            db.add(ExpenseCategory(name=cat_data["name"], description=cat_data.get("description", "")))
            imported += 1
    stats["expense_categories"] = imported

    # Import customers
    imported = 0
    for c_data in data.get("customers", []):
        existing = db.query(Customer).filter(Customer.phone_number == c_data.get("phone_number")).first()
        if not existing:
            db.add(Customer(
                customer_name=c_data.get("customer_name", ""),
                phone_number=c_data.get("phone_number", ""),
                email=c_data.get("email", ""),
                address=c_data.get("address", ""),
                gstin=c_data.get("gstin", ""),
                state=c_data.get("state", ""),
                total_purchases=c_data.get("total_purchases", 0),
            ))
            imported += 1
    stats["customers"] = imported

    # Import suppliers
    imported = 0
    for s_data in data.get("suppliers", []):
        existing = db.query(Supplier).filter(Supplier.phone_number == s_data.get("phone_number")).first()
        if not existing:
            db.add(Supplier(
                supplier_name=s_data.get("supplier_name", ""),
                contact_person=s_data.get("contact_person", ""),
                phone_number=s_data.get("phone_number", ""),
                email=s_data.get("email", ""),
                address=s_data.get("address", ""),
                gstin=s_data.get("gstin", ""),
                state=s_data.get("state", ""),
            ))
            imported += 1
    stats["suppliers"] = imported

    db.commit()
    return {"status": "success", "imported": stats}
