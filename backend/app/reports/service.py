"""Reports Service — generates Sales, Inventory, GST, Customer and P&L reports.

Each function returns a list-of-dicts ready for JSON serialization *and* for
feeding into the Excel/PDF generators.
"""

from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, ROUND_HALF_UP
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.sales.models import Sale, SaleItem, SaleStatus
from app.products.models import Product, Category
from app.customers.models import Customer
from app.inventory.models import InventoryTransaction
from app.expenses.models import Expense
from app.settings.service import get_settings


def _money(v):
    return float(Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


# ---------------------------------------------------------------------------
# 1. SALES REPORT
# ---------------------------------------------------------------------------
def sales_report(db: Session, start_date: date, end_date: date) -> dict:
    """Sales summary + line items for the given date range."""
    sales = (
        db.query(Sale)
        .filter(Sale.status == SaleStatus.completed)
        .all()
    )

    # Filter by date in Python for SQLite compatibility
    filtered = [
        s for s in sales
        if s.created_at and start_date <= s.created_at.date() <= end_date
    ]

    total_revenue = sum(float(s.total_amount or 0) for s in filtered)
    total_profit = sum(float(s.profit or 0) for s in filtered)
    total_cgst = sum(float(s.cgst or 0) for s in filtered)
    total_sgst = sum(float(s.sgst or 0) for s in filtered)
    total_igst = sum(float(s.igst or 0) for s in filtered)
    total_discount = sum(float(s.discount_amount or 0) for s in filtered)

    rows = []
    for s in filtered:
        customer = None
        if s.customer_id:
            customer = db.query(Customer).filter(Customer.id == s.customer_id).first()
        rows.append({
            "invoice_number": s.invoice_number,
            "date": s.created_at.strftime("%d/%m/%Y") if s.created_at else "",
            "customer_name": customer.customer_name if customer else "Walk-in",
            "payment_method": s.payment_method.value if hasattr(s.payment_method, 'value') else str(s.payment_method),
            "subtotal": _money(s.subtotal),
            "discount": _money(s.discount_amount),
            "cgst": _money(s.cgst),
            "sgst": _money(s.sgst),
            "igst": _money(s.igst),
            "total": _money(s.total_amount),
            "profit": _money(s.profit),
        })

    return {
        "report_type": "sales",
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "summary": {
            "total_sales": len(filtered),
            "total_revenue": _money(total_revenue),
            "total_profit": _money(total_profit),
            "total_cgst": _money(total_cgst),
            "total_sgst": _money(total_sgst),
            "total_igst": _money(total_igst),
            "total_tax": _money(total_cgst + total_sgst + total_igst),
            "total_discount": _money(total_discount),
        },
        "rows": rows,
    }


# ---------------------------------------------------------------------------
# 2. INVENTORY REPORT
# ---------------------------------------------------------------------------
def inventory_report(db: Session, stock_filter: str = "all") -> dict:
    """Current inventory snapshot — all products with stock value.

    stock_filter: 'all' | 'low' | 'out'
    """
    query = db.query(Product)
    if stock_filter == "low":
        query = query.filter(Product.stock_quantity <= Product.minimum_stock, Product.stock_quantity > 0)
    elif stock_filter == "out":
        query = query.filter(Product.stock_quantity <= 0)

    products = query.order_by(Product.product_name).all()

    rows = []
    total_stock_value = 0
    total_retail_value = 0
    for p in products:
        category = db.query(Category).filter(Category.id == p.category_id).first() if p.category_id else None
        stock_cost = float(p.purchase_price or 0) * (p.stock_quantity or 0)
        stock_retail = float(p.selling_price or 0) * (p.stock_quantity or 0)
        total_stock_value += stock_cost
        total_retail_value += stock_retail

        status = "normal"
        if (p.stock_quantity or 0) <= 0:
            status = "out_of_stock"
        elif (p.stock_quantity or 0) <= (p.minimum_stock or 0):
            status = "low"

        rows.append({
            "product_name": p.product_name,
            "category": category.name if category else "—",
            "barcode": p.barcode or "—",
            "hsn_code": p.hsn_code or "—",
            "stock_quantity": p.stock_quantity or 0,
            "minimum_stock": p.minimum_stock or 0,
            "purchase_price": _money(p.purchase_price),
            "selling_price": _money(p.selling_price),
            "stock_value": _money(stock_cost),
            "retail_value": _money(stock_retail),
            "status": status,
        })

    return {
        "report_type": "inventory",
        "stock_filter": stock_filter,
        "summary": {
            "total_products": len(rows),
            "total_stock_value": _money(total_stock_value),
            "total_retail_value": _money(total_retail_value),
            "potential_profit": _money(total_retail_value - total_stock_value),
        },
        "rows": rows,
    }


# ---------------------------------------------------------------------------
# 3. GST REPORT
# ---------------------------------------------------------------------------
def gst_report(db: Session, start_date: date, end_date: date) -> dict:
    """GST summary grouped by tax slab for the period."""
    sales = (
        db.query(Sale)
        .filter(Sale.status == SaleStatus.completed)
        .all()
    )
    filtered_sales = [
        s for s in sales
        if s.created_at and start_date <= s.created_at.date() <= end_date
    ]
    sale_ids = [s.id for s in filtered_sales]

    # Get all items for those sales
    items = db.query(SaleItem).filter(SaleItem.sale_id.in_(sale_ids)).all() if sale_ids else []

    # Group by GST slab
    slabs = {}
    for item in items:
        rate = float(item.gst_percentage or 0)
        key = f"{rate:.1f}%"
        if key not in slabs:
            slabs[key] = {"rate": rate, "taxable_value": 0, "cgst": 0, "sgst": 0, "igst": 0, "total_tax": 0, "count": 0}
        taxable = float(item.unit_price or 0) * item.quantity - float(item.discount or 0)
        slabs[key]["taxable_value"] += taxable
        slabs[key]["cgst"] += float(item.cgst or 0)
        slabs[key]["sgst"] += float(item.sgst or 0)
        slabs[key]["igst"] += float(item.igst or 0)
        slabs[key]["total_tax"] += float(item.cgst or 0) + float(item.sgst or 0) + float(item.igst or 0)
        slabs[key]["count"] += 1

    # Round values
    rows = []
    for key in sorted(slabs.keys(), key=lambda x: slabs[x]["rate"]):
        s = slabs[key]
        rows.append({
            "gst_slab": key,
            "items_count": s["count"],
            "taxable_value": _money(s["taxable_value"]),
            "cgst": _money(s["cgst"]),
            "sgst": _money(s["sgst"]),
            "igst": _money(s["igst"]),
            "total_tax": _money(s["total_tax"]),
        })

    grand_tax = sum(r["total_tax"] for r in rows)
    grand_taxable = sum(r["taxable_value"] for r in rows)

    return {
        "report_type": "gst",
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "summary": {
            "total_taxable_value": _money(grand_taxable),
            "total_cgst": _money(sum(r["cgst"] for r in rows)),
            "total_sgst": _money(sum(r["sgst"] for r in rows)),
            "total_igst": _money(sum(r["igst"] for r in rows)),
            "total_tax": _money(grand_tax),
        },
        "rows": rows,
    }


# ---------------------------------------------------------------------------
# 4. CUSTOMER REPORT
# ---------------------------------------------------------------------------
def customer_report(db: Session) -> dict:
    """All customers with purchase summary and credit info."""
    customers = db.query(Customer).order_by(Customer.customer_name).all()

    rows = []
    total_purchases = 0
    total_credit_due = 0

    for c in customers:
        # Get credit-pending sales for this customer
        pending = (
            db.query(Sale)
            .filter(
                Sale.customer_id == c.id,
                Sale.status == SaleStatus.completed,
                Sale.payment_status.in_(["partial", "unpaid"]),
            )
            .all()
        )
        credit_due = sum(float(s.amount_due or 0) for s in pending)
        sale_count = db.query(Sale).filter(Sale.customer_id == c.id, Sale.status == SaleStatus.completed).count()

        total_purchases += float(c.total_purchases or 0)
        total_credit_due += credit_due

        rows.append({
            "customer_name": c.customer_name,
            "phone_number": c.phone_number or "—",
            "email": c.email or "—",
            "state": c.state or "—",
            "gstin": c.gstin or "—",
            "total_purchases": _money(c.total_purchases or 0),
            "sale_count": sale_count,
            "credit_due": _money(credit_due),
        })

    return {
        "report_type": "customer",
        "summary": {
            "total_customers": len(rows),
            "total_purchases": _money(total_purchases),
            "total_credit_due": _money(total_credit_due),
        },
        "rows": rows,
    }


# ---------------------------------------------------------------------------
# 5. PROFIT & LOSS REPORT
# ---------------------------------------------------------------------------
def profit_loss_report(db: Session, start_date: date, end_date: date) -> dict:
    """Simple P&L: Revenue - Cost of Goods Sold - Purchase costs."""

    # --- Revenue (from completed sales) ---
    sales = db.query(Sale).filter(Sale.status == SaleStatus.completed).all()
    filtered_sales = [
        s for s in sales
        if s.created_at and start_date <= s.created_at.date() <= end_date
    ]

    revenue = sum(float(s.total_amount or 0) for s in filtered_sales)
    sales_profit = sum(float(s.profit or 0) for s in filtered_sales)
    total_discount = sum(float(s.discount_amount or 0) for s in filtered_sales)
    total_tax_collected = sum(
        float(s.cgst or 0) + float(s.sgst or 0) + float(s.igst or 0)
        for s in filtered_sales
    )

    # Cost of Goods Sold (revenue - profit - discount gives COGS approximately)
    # More accurate: sum of purchase_price * qty for each sold item
    sale_ids = [s.id for s in filtered_sales]
    items = db.query(SaleItem).filter(SaleItem.sale_id.in_(sale_ids)).all() if sale_ids else []
    cogs = 0
    for item in items:
        product = db.query(Product).filter(Product.id == item.product_id).first()
        if product:
            cogs += float(product.purchase_price or 0) * item.quantity

    # --- Purchases made in the period ---
    all_purchases = db.query(Purchase).all()
    filtered_purchases = [
        p for p in all_purchases
        if p.created_at and start_date <= p.created_at.date() <= end_date
    ]
    total_purchases = sum(float(p.total_amount or 0) for p in filtered_purchases)

    # --- Expenses in the period ---
    all_expenses = db.query(Expense).all()
    filtered_expenses = [
        e for e in all_expenses
        if start_date <= e.expense_date <= end_date
    ]
    total_expenses = sum(float(e.amount or 0) for e in filtered_expenses)

    gross_profit = revenue - cogs
    net_profit = gross_profit - total_expenses

    rows = [
        {"category": "Revenue", "description": "Total Sales Revenue", "amount": _money(revenue)},
        {"category": "Revenue", "description": "Less: Discounts Given", "amount": -_money(total_discount)},
        {"category": "Revenue", "description": "Tax Collected (GST)", "amount": _money(total_tax_collected)},
        {"category": "Cost", "description": "Cost of Goods Sold", "amount": -_money(cogs)},
        {"category": "Cost", "description": "Total Purchases (Stock Intake)", "amount": -_money(total_purchases)},
        {"category": "Cost", "description": "Operating Expenses", "amount": -_money(total_expenses)},
        {"category": "Profit", "description": "Gross Profit", "amount": _money(gross_profit)},
        {"category": "Profit", "description": "Net Profit", "amount": _money(net_profit)},
    ]

    return {
        "report_type": "profit_loss",
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "summary": {
            "total_revenue": _money(revenue),
            "total_cogs": _money(cogs),
            "total_purchases": _money(total_purchases),
            "total_expenses": _money(total_expenses),
            "gross_profit": _money(gross_profit),
            "net_profit": _money(net_profit),
            "tax_collected": _money(total_tax_collected),
            "discounts_given": _money(total_discount),
            "total_sales_count": len(filtered_sales),
        },
        "rows": rows,
    }


# ---------------------------------------------------------------------------
# EXCEL EXPORT
# ---------------------------------------------------------------------------
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def _excel_style_header(ws, row_num, max_col):
    """Apply header styling to a row."""
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border


def generate_excel(report_data: dict) -> io.BytesIO:
    """Generate an Excel workbook from report data."""
    wb = Workbook()
    ws = wb.active

    report_type = report_data.get("report_type", "report")
    ws.title = report_type.replace("_", " ").title()

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    title_cell = ws.cell(row=1, column=1, value=f"{ws.title} Report")
    title_cell.font = Font(size=14, bold=True, color="1E293B")
    title_cell.alignment = Alignment(horizontal='center')

    # Date range
    if report_data.get("start_date"):
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        ws.cell(row=2, column=1, value=f"Period: {report_data.get('start_date')} to {report_data.get('end_date')}").font = Font(size=10, color="64748B")
        ws.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    # Summary section
    summary = report_data.get("summary", {})
    if summary:
        start_row = 4
        ws.cell(row=start_row, column=1, value="Summary").font = Font(size=11, bold=True, color="4F46E5")
        row = start_row + 1
        for key, value in summary.items():
            label = key.replace("_", " ").title()
            ws.cell(row=row, column=1, value=label).font = Font(size=10, color="334155")
            ws.cell(row=row, column=2, value=value).font = Font(size=10, bold=True, color="1E293B")
            if isinstance(value, (int, float)):
                ws.cell(row=row, column=2).number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
            row += 1

    # Data rows
    rows = report_data.get("rows", [])
    if rows:
        data_start = row + 2 if summary else 4
        ws.cell(row=data_start, column=1, value="Details").font = Font(size=11, bold=True, color="4F46E5")

        # Headers
        headers = list(rows[0].keys())
        header_row = data_start + 1
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=header_row, column=col_idx, value=header.replace("_", " ").title())

        _excel_style_header(ws, header_row, len(headers))

        # Data
        alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )

        for row_idx, row_data in enumerate(rows, header_row + 1):
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))
                cell.font = Font(size=9, color="475569")
                cell.border = thin_border
                if isinstance(row_data.get(key), float):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # Auto-fit column widths (approximation)
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header.replace("_", " ").title())
            for row_data in rows[:50]:
                val = str(row_data.get(header, ""))
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# PDF EXPORT  (for reports, not invoices)
# ---------------------------------------------------------------------------
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT


def generate_report_pdf(report_data: dict, shop_settings=None) -> io.BytesIO:
    """Generate a PDF report from report data."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=12 * mm, leftMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Heading1'],
        fontSize=16, alignment=TA_CENTER, spaceAfter=2 * mm,
        textColor=rl_colors.HexColor('#1e293b'),
    )
    subtitle_style = ParagraphStyle(
        'Subtitle', parent=styles['Normal'],
        fontSize=9, alignment=TA_CENTER, textColor=rl_colors.HexColor('#64748b'),
    )

    # Shop header
    if shop_settings:
        elements.append(Paragraph(f"<b>{shop_settings.shop_name or 'Retail Store'}</b>", title_style))
        if shop_settings.shop_gstin:
            elements.append(Paragraph(f"GSTIN: {shop_settings.shop_gstin}", subtitle_style))

    # Report title
    report_type = report_data.get("report_type", "report").replace("_", " ").title()
    elements.append(Paragraph(f"<b>{report_type} Report</b>", title_style))

    if report_data.get("start_date"):
        elements.append(Paragraph(
            f"Period: {report_data['start_date']} to {report_data['end_date']}",
            subtitle_style
        ))

    elements.append(Spacer(1, 4 * mm))

    # Summary section
    summary = report_data.get("summary", {})
    if summary:
        summary_items = []
        for key, value in summary.items():
            label = key.replace("_", " ").title()
            if isinstance(value, float):
                summary_items.append(f"<b>{label}:</b> ₹{value:,.2f}")
            else:
                summary_items.append(f"<b>{label}:</b> {value}")

        summary_text = " | ".join(summary_items)
        elements.append(Paragraph(summary_text, ParagraphStyle(
            'SummaryLine', parent=styles['Normal'], fontSize=9,
            textColor=rl_colors.HexColor('#334155'), alignment=TA_CENTER,
        )))
        elements.append(Spacer(1, 5 * mm))

    # Data table
    rows = report_data.get("rows", [])
    if rows:
        headers = [h.replace("_", " ").title() for h in rows[0].keys()]
        table_data = [headers]

        for row_data in rows:
            formatted = []
            for val in row_data.values():
                if isinstance(val, float):
                    formatted.append(f"₹{val:,.2f}")
                else:
                    formatted.append(str(val))
            table_data.append(formatted)

        num_cols = len(headers)
        col_width = (doc.width) / num_cols

        table = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('TEXTCOLOR', (0, 1), (-1, -1), rl_colors.HexColor('#475569')),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#e2e8f0')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            *[('BACKGROUND', (0, i), (-1, i), rl_colors.HexColor('#f8fafc')) for i in range(2, len(table_data), 2)],
        ]))
        elements.append(table)

    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph(
        f"Generated on {datetime.now().strftime('%d/%m/%Y %I:%M %p')}",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7,
                       alignment=TA_CENTER, textColor=rl_colors.HexColor('#94a3b8'))
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer
